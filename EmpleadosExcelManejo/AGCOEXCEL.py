import tkinter as tk
from tkinter import messagebox, simpledialog
from openpyxl import load_workbook

# Cargar el archivo Excel
workbook = load_workbook(filename="empleados.xlsx")
sheet = workbook.active

# Crear una ventana de tkinter
root = tk.Tk()
root.title("Empleados registrados")

# Crear una tabla para mostrar los empleados
table = tk.Frame(root)
table.pack(padx=10, pady=10)

# Añadir encabezados a la tabla
headers = ["Nombre", "Apellido", "Edad", "Departamento"]
for col, header in enumerate(headers):
    label = tk.Label(table, text=header, font="Helvetica 12 bold", bg="lightblue", padx=10, pady=5, borderwidth=1,
                     relief="solid")
    label.grid(row=0, column=col, sticky="nsew")

# Añadir los datos de los empleados a la tabla
for row, record in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
    for col, value in enumerate(record):
        label = tk.Label(table, text=value, font="Helvetica 10", padx=10, pady=5, borderwidth=1, relief="solid")
        label.grid(row=row + 1, column=col, sticky="nsew")


# Función para borrar un empleado
def delete_employee():
    name = simpledialog.askstring("Eliminar empleado", "¿Qué empleado quieres eliminar?")
    if name is not None:
        for row, record in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            if name.lower() == record[0].lower():
                sheet.delete_rows(row + 2, amount=1)  # +2 porque iter_rows empieza en la fila 2
                workbook.save("empleados.xlsx")
                messagebox.showinfo("Empleado eliminado", f"El empleado {name} ha sido eliminado.")
                break
        else:
            messagebox.showwarning("Empleado no encontrado", f"No se ha encontrado ningún empleado llamado {name}.")

        # Eliminar todos los widgets de la tabla actual
        for widget in table.winfo_children():
            widget.destroy()

        # Añadir encabezados a la tabla
        headers = ["Nombre", "Apellido", "Edad", "Departamento"]
        for col, header in enumerate(headers):
            label = tk.Label(table, text=header, font="Helvetica 12 bold", bg="lightblue", padx=10, pady=5,
                             borderwidth=1, relief="solid")
            label.grid(row=0, column=col, sticky="nsew")

        # Añadir los datos de los empleados actualizados a la tabla
        for row, record in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            for col, value in enumerate(record):
                label = tk.Label(table, text=value, font="Helvetica 10", padx=10, pady=5, borderwidth=1, relief="solid")
                label.grid(row=row + 1, column=col, sticky="nsew")


# Botón para borrar un empleado
delete_button = tk.Button(root, text="Eliminar empleado", command=delete_employee)
delete_button.pack(padx=10, pady=10)


# Función para agregar un empleado
def add_employee():
    name = simpledialog.askstring("Agregar empleado", "Ingrese el nombre del empleado")
    last_name = simpledialog.askstring("Agregar empleado", "Ingrese el apellido del empleado")
    age = simpledialog.askinteger("Agregar empleado", "Ingrese la edad del empleado")
    department = simpledialog.askstring("Agregar empleado", "Ingrese el departamento del empleado")

    if name and last_name and age and department:  # Verificar que se hayan ingresado todos los campos
        new_employee = [name, last_name, age, department]
        sheet.append(new_employee)
        workbook.save("empleados.xlsx")
        messagebox.showinfo("Empleado agregado", f"El empleado {name} ha sido agregado.")

        # Eliminar todos los widgets de la tabla actual
        for widget in table.winfo_children():
            widget.destroy()

        # Añadir encabezados a la tabla
        headers = ["Nombre", "Apellido", "Edad", "Departamento"]
        for col, header in enumerate(headers):
            label = tk.Label(table, text=header, font="Helvetica 12 bold", bg="lightblue", padx=10, pady=5,
                             borderwidth=1, relief="solid")
            label.grid(row=0, column=col, sticky="nsew")

        # Añadir los datos de los empleados actualizados a la tabla
        for row, record in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            for col, value in enumerate(record):
                label = tk.Label(table, text=value, font="Helvetica 10", padx=10, pady=5, borderwidth=1, relief="solid")
                label.grid(row=row + 1, column=col, sticky="nsew")
    else:
        messagebox.showwarning("Campos vacíos", "Por favor complete todos los campos.")


add_button = tk.Button(root, text="Agregar empleado", command=add_employee)
add_button.pack(padx=10, pady=10)

# Ejecutar la ventana de tkinter
root.mainloop()
