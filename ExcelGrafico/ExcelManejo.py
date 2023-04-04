import openpyxl as xl
from openpyxl.chart import BarChart, Reference


# En openpyxl tenemos un modulo llamado chart
# y de este modulo estamos importando las dos clases BarChart, Reference

def process_workbook(filename):  # Funcion para modificar ese valor
    wb = xl.load_workbook(filename)  # Documento
    sheet = wb['Sheet1']  # Hoja
    # cell = sheet['a1']  # Combinaciones de la celda es lo mismo que abajo
    cell = sheet.cell(1, 1)
    # print(cell.value)

    # Iterar en cada fila de transaction
    # y por cada fila obtener el precio y después lo multiplicamos por 0.9

    # Cuántas filas tenemos en el spreed sheet
    # print(sheet.max_row) # 4

    # Crear un for que cree 1-4
    for row in range(2, sheet.max_row + 1):  # Agregamos 1 para el último index
        cell_S = sheet.cell(row, 3)
        # print(cell_S.value)  (para impimir valores de la 3ra columna)
        corrected_price = cell_S.value * 0.9  # Lo corregimos
        # Agregamos una nueva columna
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # Agregar una gráfica
    # Seleccionamos los valores de la D (3 valores)

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)  # Guardar los valores de corrected_price_cell

    chart = BarChart()  # Creamos el objeto
    chart.add_data(values)  # Agregamos la data por (valores)
    sheet.add_chart(chart, 'e2')  # Ponemos los valores en la E2

    wb.save(filename)  # Guardarlo en caso de que nuestro programa tenga un bug
